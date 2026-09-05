import io
import csv
from datetime import datetime
from decimal import Decimal

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_budget_to_pdf(budget_items, categories, report_period='Mensuel', academic_year='2024-2025'):
    if not REPORTLAB_AVAILABLE:
        raise ImportError('reportlab is not installed')
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1A3A36'),
        spaceAfter=12,
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#5F726D'),
        spaceAfter=20,
    )
    
    story = []
    story.append(Paragraph('Rapport Budgétaire', title_style))
    story.append(Paragraph(f'{report_period} · {academic_year} · Généré le {datetime.now().strftime("%d/%m/%Y %H:%M")}', subtitle_style))
    story.append(Spacer(1, 0.5*cm))
    
    total_revenue = sum(Decimal(str(item.amount)) for item in budget_items if item.item_type == 'REVENUE')
    total_expense = sum(Decimal(str(item.amount)) for item in budget_items if item.item_type == 'EXPENSE')
    balance = total_revenue - total_expense
    
    summary_data = [
        ['Indicateur', 'Montant (Ar)'],
        ['Total Recettes', f'{total_revenue:,.2f}'.replace(',', ' ')],
        ['Total Dépenses', f'{total_expense:,.2f}'.replace(',', ' ')],
        ['Solde', f'{balance:,.2f}'.replace(',', ' ')],
    ]
    
    summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A3A36')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F4F2')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E8E4')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 1*cm))
    
    if budget_items:
        items_data = [['Date', 'Type', 'Catégorie', 'Description', 'Montant (Ar)']]
        for item in budget_items[:100]:
            cat_name = categories.get(item.category, {}).get('name', '—') if isinstance(categories, dict) else '—'
            items_data.append([
                item.date,
                'Recette' if item.item_type == 'REVENUE' else 'Dépense',
                cat_name,
                item.description[:50],
                f'{Decimal(str(item.amount)):,.2f}'.replace(',', ' '),
            ])
        
        items_table = Table(items_data, colWidths=[2.5*cm, 2.5*cm, 3*cm, 5*cm, 3*cm])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2B6F68')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAF9')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E8E4')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        story.append(items_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def export_budget_to_excel(budget_items, categories, report_period='Mensuel', academic_year='2024-2025'):
    if not OPENPYXL_AVAILABLE:
        raise ImportError('openpyxl is not installed')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Budget'
    
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1A3A36', end_color='1A3A36', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin', color='E0E8E4'),
        right=Side(style='thin', color='E0E8E4'),
        top=Side(style='thin', color='E0E8E4'),
        bottom=Side(style='thin', color='E0E8E4'),
    )
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f'Rapport Budgétaire - {report_period} {academic_year}'
    ws['A1'].font = Font(bold=True, size=14, color='1A3A36')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='E8F0EC', end_color='E8F0EC', fill_type='solid')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Généré le {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10, color='5F726D')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['Date', 'Type', 'Catégorie', 'Description', 'Désignation', 'Montant (Ar)']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    total_revenue = Decimal('0')
    total_expense = Decimal('0')
    
    for row_num, item in enumerate(budget_items[:500], start=5):
        cat_name = categories.get(item.category, {}).get('name', '—') if isinstance(categories, dict) else '—'
        amount = Decimal(str(item.amount))
        
        if item.item_type == 'REVENUE':
            total_revenue += amount
        else:
            total_expense += amount
        
        ws.cell(row=row_num, column=1, value=item.date).alignment = cell_alignment
        ws.cell(row=row_num, column=2, value='Recette' if item.item_type == 'REVENUE' else 'Dépense').alignment = cell_alignment
        ws.cell(row=row_num, column=3, value=cat_name).alignment = cell_alignment
        ws.cell(row=row_num, column=4, value=item.description).alignment = cell_alignment
        ws.cell(row=row_num, column=5, value=item.designation or '').alignment = cell_alignment
        ws.cell(row=row_num, column=6, value=float(amount)).alignment = cell_alignment
        ws.cell(row=row_num, column=6).number_format = '#,##0.00'
        
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = border
    
    summary_row = len(budget_items[:500]) + 6
    ws.cell(row=summary_row, column=1, value='Total Recettes').font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=float(total_revenue)).font = Font(bold=True)
    ws.cell(row=summary_row, column=6).number_format = '#,##0.00'
    
    ws.cell(row=summary_row + 1, column=1, value='Total Dépenses').font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=6, value=float(total_expense)).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=6).number_format = '#,##0.00'
    
    ws.cell(row=summary_row + 2, column=1, value='Solde').font = Font(bold=True, color='1A3A36')
    ws.cell(row=summary_row + 2, column=6, value=float(total_revenue - total_expense)).font = Font(bold=True, color='1A3A36')
    ws.cell(row=summary_row + 2, column=6).number_format = '#,##0.00'
    
    for col in range(1, 7):
        for row in range(summary_row, summary_row + 3):
            ws.cell(row=row, column=col).border = border
    
    for col in ws.columns:
        max_length = 0
        column = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                column = cell.column_letter
                break
        if column is None:
            continue
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_budget_to_csv(budget_items, categories, report_period='Mensuel', academic_year='2024-2025'):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    
    writer.writerow(['Rapport Budgétaire', report_period, academic_year])
    writer.writerow(['Généré le', datetime.now().strftime('%d/%m/%Y %H:%M'), ''])
    writer.writerow([])
    writer.writerow(['Date', 'Type', 'Catégorie', 'Description', 'Désignation', 'Montant (Ar)'])
    
    total_revenue = Decimal('0')
    total_expense = Decimal('0')
    
    for item in budget_items[:1000]:
        cat_name = categories.get(item.category, {}).get('name', '—') if isinstance(categories, dict) else '—'
        amount = Decimal(str(item.amount))
        
        if item.item_type == 'REVENUE':
            total_revenue += amount
        else:
            total_expense += amount
        
        writer.writerow([
            item.date,
            'Recette' if item.item_type == 'REVENUE' else 'Dépense',
            cat_name,
            item.description,
            item.designation or '',
            f'{amount:,.2f}'.replace(',', ' '),
        ])
    
    writer.writerow([])
    writer.writerow(['Total Recettes', f'{total_revenue:,.2f}'.replace(',', ' '), ''])
    writer.writerow(['Total Dépenses', f'{total_expense:,.2f}'.replace(',', ' '), ''])
    writer.writerow(['Solde', f'{total_revenue - total_expense:,.2f}'.replace(',', ' '), ''])
    
    buffer.seek(0)
    return buffer